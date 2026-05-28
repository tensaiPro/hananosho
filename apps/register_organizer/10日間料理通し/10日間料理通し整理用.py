import sys
import os
import tkinter.messagebox as msgbox
import win32com.client
import datetime

import yaml
import openpyxl as opxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment, colors
from openpyxl.cell.cell import MergedCell

#import pandas as pd

# yamlファイル読み込み
with open("10日間料理通し.yml", "r", encoding = "utf-8") as yml:
    config = yaml.safe_load(yml)

file_name = os.getcwd() + "\\" + config["xls_file_name"]
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False

# ファイルがあれば開きなければエラーメッセージを表示してプログラム終了
if(os.path.isfile(file_name)):
    wb = excel.Workbooks.Open(file_name)
    file_name = os.getcwd() + "\\" + config["xlsx_file_name"]
    wb.SaveAs(file_name, FileFormat=51) #51 = xlsx
    wb.close
else:
    file_name = os.getcwd() + "\\" + config["xlsx_file_name"]
    if(os.path.isfile(file_name)):
        msgbox.info("確認", f"「{config["xls_file_name"]}」がないため「{config["xlsx_file_name"]}」を参照します。")
    else:
        msgbox.showerror("エラー", f"{config["xls_file_name"]}、または{config["xlsx_file_name"]}がありません")

#file_name = os.getcwd() + "\\" + config["xlsx_file_name"]


wb_source = Workbook()
file_name = config["xlsx_file_name"]
if(os.path.isfile(file_name)):
    wb_source = load_workbook(file_name)
else:
    msgbox.showerror("エラー", f"{file_name}がありません")
    sys.exit(0)

ws = wb_source.active

target_range = CellRange(f"{config["option"]["sColumn"]}{config["option"]["sRow"]}:{config["option"]["eColumn"]}{config["option"]["eRow"]}")

if config["output"]["new_excel_file"]:
    wb_new = Workbook()
    ws_new = wb_new.worksheets[0]
    ws_new.title = f"{config["output"]["sheet_name"]}"
else:
    ws_new = wb_source.create_sheet(title=f"{config["output"]["sheet_name"]}")

#別注料理からの行以下のセルを結合解除
for merged_range in list(ws.merged_cells.ranges):
    # 3. ターゲット範囲と結合範囲が重なっているか判定
    # ※少しでも重なれば解除する場合は「issubset」ではなく「intersect」などを使う
    if target_range.issubset(merged_range) or merged_range.issubset(target_range) or \
       (set(target_range.cells) & set(merged_range.cells)):
        ws.unmerge_cells(str(merged_range))

#別注料理からの行以下のセルのValueを削除、書式をクリア
for row in range(config["option"]["sRow"] -1 , config["option"]["eRow"], 1):
    for column in range(column_index_from_string(config["option"]["sColumn"]),column_index_from_string(config["option"]["eColumn"]), 1):
        cell = ws[f"{get_column_letter(column)}{row}"]
        cell.value = None
        #cell.font = Font()           # フォントを初期化
        cell.fill = PatternFill(fill_type=None) # 塗りつぶしを解除
        cell.border = Border()       # 罫線を解除
        cell.alignment = Alignment() # 配置を初期化
        cell.number_format = 'General' # 表示形式を標準に

# 料理ランクがいくつあるかを調べる
#ws = wb_source.worksheets[0]
rank_row_end = config["rank"]["sRow"]
while ws.cell(column = config["rank"]["sColumn"], row = rank_row_end).value != config["rank"]["endRowValue"]:
    rank_row_end += 1

# 日ごとのまとめ表の行のスタート位置
table_sRow = rank_row_end + config["table"]["space_from_rank"]

#全日のランクの名前とランクごとの人数を格納するリスト
alldays_rank_name_and_number = []
# 日ごとのランク毎の人数を収納するリスト
rank_name_and_number = []
# 別注料理の行スタート位置
optPos = config["option"]["sRow"]

#日ごとのランクを集計して日付と一緒にリストに格納
for day in range(config["date"]["start_column"],\
                    config["date"]["start_column"] + config["span"]["days"] * config["span"]["columns_each_day"],\
                    config["span"]["columns_each_day"]):
    rank_name_and_number.clear()
    rank_name_and_number.append(ws.cell(row=config["date"]["row"],column=day).value)
    for rank in range(config["rank"]["sRow"], rank_row_end):
        rank_count = 0
        for adult_and_child in range(day, day + config["span"]["columns_each_day"] - config["guest_type"]["merged_column_childB"]):
            rank_count += int(ws.cell(row = rank, column = adult_and_child).value)
            if ws.cell(row = rank, column = adult_and_child).value == 0:
                ws.cell(row = rank, column = adult_and_child).value = None
        else:
            if rank_count:
                rank_name_and_number.append([ws.cell(row = rank, column = config["rank"]["sColumn"]).value, rank_count])
    else:
        # msgbox.showinfo("", rank_name_and_number) # 日ごとの料理ランクを集計したらメッセージボックスで表示
        alldays_rank_name_and_number.append(rank_name_and_number.copy())
        for i in range(len(rank_name_and_number)):
            if not i:
                ws.cell(row = table_sRow+i, column = day).value = rank_name_and_number[i]
            else:
                ws.cell(row = table_sRow+i, column = day).value = rank_name_and_number[i][0]

            if i:
                ws.cell(row = table_sRow+i, column = day+1).value = rank_name_and_number[i][1]

#10日間を何分割するか決定
days_divided_by = 0
if config["new_sheet"]["general"]["days_divided_by"] >= 2 and config["new_sheet"]["general"]["days_divided_by"] <= 10:
    days_divided_by = config["new_sheet"]["general"]["days_divided_by"]
elif config["new_sheet"]["general"]["days_divided_by"] > 10:
    days_divided_by = 10
elif config["new_sheet"]["general"]["days_divided_by"] < 2:
    days_divided_by = 2
else:
    msgbox.showerror("エラー", "config[""new_sheet""][""general""][""days_divided_by""]の値が不正です。数値を入力してください。")
    sys.exit(0)

#Aチャメを残すか判定とその場合の列幅、列数を計算
if config["new_sheet"]["column_settings"]["del_childA"]:
    NEW_COLUMN_EACH_DAYS = 2
else:
    NEW_COLUMN_EACH_DAYS = 3

if NEW_COLUMN_EACH_DAYS == 2:
    NEW_DAYS_COLUMN_WIDTH = config["new_sheet"]["column_width"]["adult_and_child"] * 3 / 2
else:
    NEW_DAYS_COLUMN_WIDTH = config["new_sheet"]["column_width"]["adult_and_child"]

#分割したそれぞれの行列の情報を格納する変数
part_start_row = [0] * days_divided_by
part_end_row = [0] * days_divided_by

day_divmod = divmod(config["new_sheet"]["span"]["days"], days_divided_by)

#ここから新しいシートに入力
for part in range(days_divided_by):
    space = config["new_sheet"]["row_settings"]["space_rows"] if part else config["new_sheet"]["row_settings"]["start_days"]
    part_start_row[part] = space
    if part:
        part_start_row[part] += part_end_row[part - 1]
    #ランクの列を入力、とりあえず全コピー
    for row in range(rank_row_end - config["date"]["row"] + 1):
        ws_new.cell(row = part_start_row[part] + row, column = config["new_sheet"]["column_settings"]["rank_column"]).value = \
        ws.cell(row = config["date"]["row"] + row, column = config["rank"]["sColumn"]).value
    else:
        part_end_row[part] = part_start_row[part] + row

    #日毎の大人、小人をコピー
    for day in range(day_divmod[0] + (1 if part < day_divmod[1] else 0)):
        for row in range(rank_row_end - config["date"]["row"] + 1):
            for adult_and_child in range(config["guest_type"]["types"]):

                skip_childA = 0
                mod_sub = 0
                if config["new_sheet"]["column_settings"]["del_childA"] and adult_and_child >= config["guest_type"]["childA"]:
                    if row == config["date"]["row_guest_type"] - 1 and adult_and_child == config["guest_type"]["childA"]:
                        continue
                    skip_childA = 1

                if day_divmod[1]:
                    mod_sub = part
                    if day_divmod[1] < part:
                        mod_sub = day_divmod[1]
                #コピー先シートの行と列
                new_target_row = part_start_row[part] + row
                new_target_column = config["new_sheet"]["column_settings"]["start_days"] + day * NEW_COLUMN_EACH_DAYS + \
                adult_and_child - skip_childA
                #コピー元シートの行と列
                source_target_row = config["date"]["row"] + row
                source_target_column = config["date"]["start_column"] + \
                (((part * day_divmod[0]) + day) + mod_sub) * config["span"]["columns_each_day"] + \
                + adult_and_child
                #print(f"part = {part}, day = {day}, row = {row}, adult_and_child = {adult_and_child}")
                #print(f"new_sheet = {ws_new.cell(row = new_target_row, column = new_target_column).coordinate}, new_sheet = {ws.cell(row = source_target_row, column = source_target_row).coordinate}")

                if isinstance(ws.cell(row = source_target_row, \
                        column = source_target_column).value, (int, str)):
                    if isinstance(ws_new.cell(row = new_target_row, column = new_target_column).value, int):
                        ws_new.cell(row = new_target_row, \
                            column = new_target_column).value += \
                        ws.cell(row = source_target_row, \
                            column = source_target_column).value
                    else:
                        ws_new.cell(row = new_target_row, \
                            column = new_target_column).value = \
                        ws.cell(row = source_target_row, \
                            column = source_target_column).value

                else:
                    continue

    #入力した分の不要な行を削除
    now_row = part_start_row[part] + config["new_sheet"]["row_settings"]["rows_day_and_guest_type"]
    for row in range(part_end_row[part] - part_start_row[part] - config["new_sheet"]["row_settings"]["rows_day_and_guest_type"]):
        for column in range((day_divmod[0] + (1 if part < day_divmod[1] else 0)) * NEW_COLUMN_EACH_DAYS):
            if ws_new.cell(row = now_row,
                           column = config["new_sheet"]["column_settings"]["start_days"] + column).value:
                now_row += 1
                break
        else:
            ws_new.delete_rows(now_row)
            for rows in range(part, len(part_start_row)):
                part_start_row[rows] -= (0 if part == rows else 1)
                part_end_row[rows] -= 1

part_summary_start = [0] * days_divided_by
part_summary_end = [0] * days_divided_by

#集計結果を入力
for part in range(days_divided_by):
    part_summary_start[part] = part_end_row[part] + config["new_sheet"]["row_settings"]["space_rows"]
    for summary in range(day_divmod[0] + (1 if part < day_divmod[1] else 0)):
        if part_summary_end[part] < len(alldays_rank_name_and_number[part * day_divmod[0] + summary]):
            part_summary_end[part] = len(alldays_rank_name_and_number[part * day_divmod[0] + summary])
    else:
        part_summary_end[part] += part_summary_start[part] + config["new_sheet"]["row_settings"]["space_rows"] - 1
        summary_sub = part_summary_end[part] - part_summary_start[part]
        ws_new.insert_rows(part_summary_start[part], summary_sub)
        for rows in range(part, len(part_start_row)):
            part_end_row[rows] += summary_sub
            if rows == part:
                continue
            part_start_row[rows] += summary_sub

    for day in range(day_divmod[0] + (1 if part < day_divmod[1] else 0)):
        mod_sub = 0
        if day_divmod[1]:
            mod_sub = part
            if day_divmod[1] < part:
                mod_sub = day_divmod[1]
        for row, rank in enumerate(alldays_rank_name_and_number[part * day_divmod[0] + mod_sub + day]):
            column = config["new_sheet"]["column_settings"]["start_days"] + day * NEW_COLUMN_EACH_DAYS
            if not row:
                ws_new.cell(row = part_summary_start[part] + row, column = column+config["new_sheet"]["general"]["column_combined_summary_rank_and_number"]).value = \
                rank
                continue
            current_day = part * day_divmod[0] + day + mod_sub
            if config["new_sheet"]["general"]["combine_summary_rank_and_number"]:
                ws_new.cell(row = part_summary_start[part] + row, column = column+config["new_sheet"]["general"]["column_combined_summary_rank_and_number"]).value = \
                rank[0]
                ws_new.cell(row = part_summary_start[part] + row, column = column+config["new_sheet"]["general"]["column_combined_summary_rank_and_number"]).value += \
                (" " * config["new_sheet"]["general"]["spaces_combined_summary_rank_and_number"]) + str(rank[1])
                continue
            ws_new.cell(row = part_summary_start[part] + row, column = column).value = \
            rank[0]
            ws_new.cell(row = part_summary_start[part] + row, column = column + 1).value = \
            rank[1]


#新しいシートの設定開始位置

#倍率設定
ws_new.sheet_view.zoomScale = config["new_sheet"]["general"]["magnification"]
ws_new.sheet_view.zoomScaleNormal = config["new_sheet"]["general"]["magnification"]

#フォント、色、罫線の設定
font_name = config["new_sheet"]["fonts"]["font"]
#塗りつぶし設定
for part in range(days_divided_by):
    #罫線を作成
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    for column in range(config["new_sheet"]["column_settings"]["rank_column"],
                        config["new_sheet"]["column_settings"]["rank_column"] + \
                        (day_divmod[0] + (1 if part < day_divmod[1] else 0)) * NEW_COLUMN_EACH_DAYS + 1):
        for rank in range(part_start_row[part], part_summary_start[part] - config["new_sheet"]["row_settings"]["space_rows"] + 1):
            ws_new.cell(row = rank, column = column).border = border
    font_sum = Font(name=font_name, size = config["new_sheet"]["fonts"]["size_sum_row"])
    for column in range(config["new_sheet"]["column_settings"]["rank_column"],
                        config["new_sheet"]["column_settings"]["rank_column"] + \
                        NEW_COLUMN_EACH_DAYS * (day_divmod[0] + (1 if part < day_divmod[1] else 0))+1):
        #合計の行に塗りつぶしとフォント設定
        ws_new.cell(row = part_end_row[part] - (part_summary_end[part] - part_summary_start[part]), column = column).fill = \
        PatternFill(patternType="solid", fgColor = config["new_sheet"]["fonts"]["sum_row_color"])
        ws_new.cell(row = part_end_row[part] - (part_summary_end[part] - part_summary_start[part]), column = column).font = font_sum
        ws_new.cell(row = part_end_row[part] - (part_summary_end[part] - part_summary_start[part]), column = column)
        #料理、日付と人数種別に塗りつぶし
        for dish_date_guest_type in range(config["new_sheet"]["row_settings"]["rows_day_and_guest_type"]):
            ws_new.cell(row = part_start_row[part] + dish_date_guest_type, column = column).fill = \
            PatternFill(patternType="solid", fgColor = colors.Color(indexed=config["new_sheet"]["fonts"]["dish_row_fgColor_indexed"]))
    #まとめの罫線を作成
    border_left = Border(left=side)
    border_bottom = Border(bottom=side)
    border_right = Border(right=side)

    for day in range(day_divmod[0] + (1 if part < day_divmod[1] else 0)):
        for row in range(part_summary_start[part] - config["new_sheet"]["row_settings"]["space_rows"] + 1, part_end_row[part] + 1):
            ws_new.cell(row = row, column = config["new_sheet"]["column_settings"]["start_days"] + day * NEW_COLUMN_EACH_DAYS).border = border_left
        else:
            for last_row_column in range(NEW_COLUMN_EACH_DAYS):
                if not last_row_column:
                    border_leftbottom = border_left + border_bottom
                    ws_new.cell(row = row, \
                                column = config["new_sheet"]["column_settings"]["start_days"] + \
                                day * NEW_COLUMN_EACH_DAYS + last_row_column).border = border_leftbottom
                    continue
                ws_new.cell(row = row, \
                            column = config["new_sheet"]["column_settings"]["start_days"] + \
                            day * NEW_COLUMN_EACH_DAYS + last_row_column).border = border_bottom
    else:
        for row in range(part_summary_start[part] - config["new_sheet"]["row_settings"]["space_rows"] + 1, part_end_row[part] + 1):
            ws_new.cell(row = row, column = config["new_sheet"]["column_settings"]["start_days"] + day * NEW_COLUMN_EACH_DAYS + NEW_COLUMN_EACH_DAYS - 1).border = border_right
        else:
            border_rightbottom = border_bottom + border_right
            ws_new.cell(row = row, column = config["new_sheet"]["column_settings"]["start_days"] + day * NEW_COLUMN_EACH_DAYS + NEW_COLUMN_EACH_DAYS - 1).border = border_rightbottom
    #罫線作成終わり

    #料理、ランク名のセルを設定(合計のセルは別処理)
    font_dish = Font(name=font_name, size = config["new_sheet"]["fonts"]["size_dish"])
    font_rank_name = Font(name=font_name, size = config["new_sheet"]["fonts"]["size_rank_name"])
    for dish_to_rank in range(part_start_row[part], part_end_row[part] - (part_summary_end[part] - part_summary_start[part])):
        if dish_to_rank == part_start_row[part]:
            ws_new.cell(row = dish_to_rank, column = config["new_sheet"]["column_settings"]["rank_column"]).font = font_dish
            ws_new.cell(row = dish_to_rank, column = config["new_sheet"]["column_settings"]["rank_column"]).alignment = Alignment(horizontal = "center", vertical = "center")
            ws_new.merge_cells(start_row = dish_to_rank, start_column = config["new_sheet"]["column_settings"]["rank_column"], \
                            end_row = dish_to_rank + 1, end_column = config["new_sheet"]["column_settings"]["rank_column"])
            continue
        if dish_to_rank == part_start_row[part] + 1:
            continue
        ws_new.cell(row = dish_to_rank, column = config["new_sheet"]["column_settings"]["rank_column"]).font = font_rank_name
        ws_new.cell(row = dish_to_rank, column = config["new_sheet"]["column_settings"]["rank_column"]).alignment = Alignment(horizontal = "left", vertical = "center")


    #日付、人数種別、人数のセルを設定
    font_day = Font(name=font_name, size = config["new_sheet"]["fonts"]["size_day"])
    font_adult_and_child = Font(name=font_name, size = config["new_sheet"]["fonts"]["size_guest_type"])
    font_number = Font(name=font_name, size = config["new_sheet"]["fonts"]["size_number"])
    font_summary_date = Font(name=font_name, size = config["new_sheet"]["fonts"]["size_summary"], u = config["new_sheet"]["general"]["summary_date_under_line"])
    font_summary = Font(name=font_name, size = config["new_sheet"]["fonts"]["size_summary"], u = config["new_sheet"]["general"]["summary_under_line"])

    for day in range(day_divmod[0] + (1 if part < day_divmod[1] else 0)):
        #日付の設定
        ws_new.cell(row = part_start_row[part], column = config["new_sheet"]["column_settings"]["start_days"] + \
                    day * NEW_COLUMN_EACH_DAYS).font = font_day
        ws_new.cell(row = part_start_row[part], column = config["new_sheet"]["column_settings"]["start_days"] + \
                    day * NEW_COLUMN_EACH_DAYS).alignment = Alignment(horizontal = "center", vertical = "center")
        ws_new.merge_cells(start_row = part_start_row[part], end_row = part_start_row[part], \
                           start_column = config["new_sheet"]["column_settings"]["start_days"] + day * NEW_COLUMN_EACH_DAYS, \
                           end_column = config["new_sheet"]["column_settings"]["start_days"] + day * NEW_COLUMN_EACH_DAYS + NEW_COLUMN_EACH_DAYS - 1)
        for row in range(part_start_row[part] + 1, part_end_row[part] - (part_summary_end[part] - part_summary_start[part]) + 1):
            for adult_and_child in range(NEW_COLUMN_EACH_DAYS):
                column = config["new_sheet"]["column_settings"]["start_days"] + NEW_COLUMN_EACH_DAYS * day + adult_and_child
                if row == part_start_row[part] + 1:
                    ws_new.cell(row = row, column = column).font = font_adult_and_child
                    ws_new.cell(row = row, column = column).alignment = Alignment(horizontal = "center", vertical = "center")
                    continue
                if row == part_end_row[part] - (part_summary_end[part] - part_summary_start[part]):
                    ws_new.cell(row = row, column = column).font = font_sum
                    ws_new.cell(row = row, column = column).alignment = Alignment(horizontal = "right", vertical = "center")
                    continue
                ws_new.cell(row = row, column = column).font = font_number
                ws_new.cell(row = row, column = column).alignment = Alignment(horizontal = "right", vertical = "center")



        #集計結果の設定
        mod_sub = 0
        if day_divmod[1]:
            mod_sub = part
            if day_divmod[1] < part:
                mod_sub = day_divmod[1]
        today = day_divmod[0] * part + day + mod_sub
        for i in range(len(alldays_rank_name_and_number[today])):
            #print(part_summary_start[part] + i)
            column = day * NEW_COLUMN_EACH_DAYS + config["new_sheet"]["general"]["column_combined_summary_rank_and_number"] + config["new_sheet"]["column_settings"]["start_days"]
            ws_new.cell(row = part_summary_start[part] + i, column = column).font = font_summary if i else font_summary_date
            ws_new.cell(row = part_summary_start[part] + i, column = column).alignment = Alignment(horizontal = "right", vertical = "center")

ws_new.merge_cells(start_row = config["new_sheet"]["row_settings"]["title_start"], \
                   end_row = config["new_sheet"]["row_settings"]["title_end"], \
                   start_column = config["new_sheet"]["column_settings"]["title_start"], \
                   end_column = config["new_sheet"]["column_settings"]["title_end"])

#タイトル(日付+10日間料理通し)の文字列が入るセルの設定
today = datetime.datetime.today()
month = today.month
day = today.day
weekdays = ["月", "火", "水", "木", "金", "土", "日"]
weekday = weekdays[today.weekday()]
ws_new.cell(row=config["new_sheet"]["row_settings"]["title_start"], \
            column = config["new_sheet"]["column_settings"]["title_start"]).value = f"{month}/{day}({weekday}) 10日間料理通し"
font_title = Font(name=font_name, size = config["new_sheet"]["fonts"]["size_title"])
ws_new.cell(row=config["new_sheet"]["row_settings"]["title_start"], \
            column = config["new_sheet"]["column_settings"]["title_start"]).font = font_title
ws_new.cell(row=config["new_sheet"]["row_settings"]["title_start"], \
            column = config["new_sheet"]["column_settings"]["title_start"]).alignment = Alignment(horizontal = "center", vertical = "center")

#新しいシートの列幅設定
ws_new.column_dimensions[get_column_letter(1)].width = config["new_sheet"]["column_width"]["rank"]
for column in range(config["new_sheet"]["column_settings"]["start_days"], \
                    config["new_sheet"]["column_settings"]["start_days"] + config["new_sheet"]["span"]["days"] * NEW_COLUMN_EACH_DAYS,\
                    NEW_COLUMN_EACH_DAYS):
    for adult_and_child in range(NEW_COLUMN_EACH_DAYS):
        ws_new.column_dimensions[get_column_letter(column+adult_and_child)].width = NEW_DAYS_COLUMN_WIDTH

#新しいシートの行の高さ設定

#新しいシートの印刷設定
ws_new.page_setup.paperSize = ws_new.PAPERSIZE_A3
ws_new.page_setup.fitToPage = True
# 横1ページに収める
ws_new.page_setup.fitToWidth = 1
# 縦1ページに収める (必要に応じて調整)
ws_new.page_setup.fitToHeight = 1
ws_new.page_setup.orientation = ws_new.ORIENTATION_LANDSCAPE


#ここで新しいシートの入力終了

# ワークブックの変更を保存、もしくはテンプレートブックから別名保存
#if config["output_oldstyle"]:
wb_source.save(config["xlsx_file_name"])
# ワークブックを閉じる
wb_source.close()

#エクセルファイルを新規作成する場合の保存とファイルクローズ
if config["output"]["new_excel_file"]:
    wb_new.save(config["new_file_name"])
    wb_new.close()
