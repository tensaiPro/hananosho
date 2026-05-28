import sys
import os

import tkinter.messagebox as msgbox
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import yaml

wb = Workbook()

filename = "フロント台帳.xlsx"
sheetname = "fixed"

# ファイルがあれば開きなければエラーメッセージを表示してプログラム終了
if(os.path.isfile(filename)):
    wb = load_workbook(filename)
else:
    msgbox.showerror("エラー", f"{filename}がありません")
    sys.exit(0)

with open("フロント台帳.yml", "r") as yml:
    config = yaml.safe_load(yml)

colStrtBlwC = config["cells"]["sUnrqurdPplClm"]
colEndBlwC = config["cells"]["eUnrqurdPplClm"]
sGuestname = config["guestname"]["sRow"]
clmnAV = config["columnAV"]

ws_orig = wb.worksheets[0]
#ws.cell(column = colStrtBlwC-1, row = 1).value = "ここより右"
#ws.cell(column = colEndBlwC+1 , row = 1).value = "ここより左"
#ws.delete_cols(colStrtBlwC, 12)
exist = False
for ws in wb.worksheets:
    msgbox.showinfo("", ws.title)
    if ws.title is sheetname:
        exist = True
            break

# シートがあれば選択、なければ作成
if exist:
    ws_new = wb.worksheets[sheetname]
else:
    ws_new = wb.create_sheet(title = sheetname)


# ws_new.cell()

#cells = []

ws_orig.delete_rows(50) # これだと改ページが固定になる
# print(f"改ページは{ws_orig.row_breaks}")

ws_new.column_dimensions["AV"].width = 25

#for row in ws["A1:BL9"]:
#    for cell in row:
#        if type(cell) is openpyxl.cell.cell.MergedCell:
#            ws.unmerge_cells(start_row = openpyxl.coordinate)

#msgbox.showinfo("", cells)

wb.save(filename)
wb.close()
