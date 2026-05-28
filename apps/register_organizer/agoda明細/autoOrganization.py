import sys
import os
import tkinter.messagebox as msgbox

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side

# yamlファイル読み込み
with open("material_names.yml", "r", encoding='utf-8') as yml:
    config = yaml.safe_load(yml)

file_name = config['csv_file_name']

#csvのファイルを取得
if(os.path.isfile(file_name)):
    df = pd.read_csv(file_name, index_col=False)
else:
    msgbox.showerror("エラー", f"{file_name}がありません")
    sys.exit(0)

#csvをxlsxとして保存
file_name = config['excel_file_name']
df.to_excel(file_name, index=False)

#ワークブックのインスタンス化
wb = Workbook()

# ファイルがあれば開きなければエラーメッセージを表示してプログラム終了
if(os.path.isfile(file_name)):
    wb = load_workbook(file_name)
else:
    msgbox.showerror("エラー", f"{file_name}がありません")
    sys.exit(0)

delete_cols = config["delete_cols"]

#シートを選択し、不要な列を削除
ws = wb.worksheets[0]
for col in delete_cols:
    ws.delete_cols(col)

column_width = config["column_width"]
for width in column_width:
    ws.column_dimensions[width[1]].width = width[0]

#何行あるかを計算
max_row = 0
for i, row in enumerate(ws["A"]):
    #print(row.value)
    max_row += 1
    if row.value == None:
        break

#取引額0の行を削除
current_row = 0
for cell in ws["C"]:
    current_row += 1
    if cell.value == None:
        break
    if cell.value == 0:
        ws.delete_rows(current_row)
        max_row -= 1
        current_row -= 1

#不要な行を削除
res_num = 0
while res_num < max_row:
    res_num += 1
    for other_res_num in range(res_num + 1, max_row + 1, 1):
        if ws[f"A{res_num}"].value == ws[f"A{other_res_num}"].value:
            if ws[f"C{res_num}"].value + ws[f"C{other_res_num}"].value == 0:

                ws.delete_rows(other_res_num)
                ws.delete_rows(res_num)
                max_row -= 2
                res_num -= 1
                continue

#格子の枠を各セルに作成
side = Side(style="thin", color = "000000")
border = Border(top=side, bottom=side, left=side, right=side)

max_column = 1

while ws.cell(row=1, column=max_column).value != None:
    max_column += 1

for current_row in range(1, max_row + 1, 1):
    for current_column in range(1, max_column, 1):
        ws.cell(row=current_row, column=current_column).border = border

#集計用の式を挿入
ws[f"C{max_row + 1}"] = f"=SUM(C2:C{max_row})"

wb.save(file_name)

wb.close()
